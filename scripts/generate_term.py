"""Generate a new term template from the current term data."""

import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN_BORDER = Border(bottom=Side(style="thin", color="E2E8F0"))

LESSON_NAMES = [
    "Meet the Robot",
    "Power Up & Roll",
    "Robot Choreography",
    "Give It Senses",
    "The Grand Challenge",
]
LESSON_WEEKS = ["Wks 1-2", "Wks 3-4", "Wks 5-6", "Wks 7-8", "Wks 9-10"]


def generate_term(
    classes_source: str,
    term_name: str,
    output_path: str,
):
    with open(classes_source) as f:
        data = json.load(f)

    classes = data.get("classes", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lesson Tracker"

    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = f"{term_name} — Coding & Robotics Lesson Tracker"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = "GigoToys S4A Robotics  |  Eshowe Junior School  |  Mr Dlamini  |  Mark each cell: Started or Done"
    sub.font = Font(size=10)
    sub.alignment = Alignment(horizontal="center")

    headers = ["Day", "Period", "Grade", "Class"] + [
        f"Lesson {i+1}\n({LESSON_WEEKS[i]})\n{LESSON_NAMES[i]}" for i in range(5)
    ] + ["% Complete", "Notes"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[4].height = 60

    for row_idx, cls in enumerate(classes, 5):
        ws.cell(row=row_idx, column=1, value=cls["day"])
        ws.cell(row=row_idx, column=2, value=cls["period"])
        ws.cell(row=row_idx, column=3, value=cls["grade"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=4, value=cls["classCode"]).font = Font(bold=True, color="1F3864")

        for i in range(5):
            ws.cell(row=row_idx, column=5 + i).alignment = Alignment(horizontal="center")

        pct = ws.cell(row=row_idx, column=10)
        pct.value = f'=COUNTIF(E{row_idx}:I{row_idx},"Done")/5'
        pct.number_format = "0%"
        pct.font = Font(bold=True)
        pct.alignment = Alignment(horizontal="center")

        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    summary_row = 5 + len(classes) + 1
    ws.cell(row=summary_row, column=1, value="Classes Done — per lesson").font = Font(bold=True)
    for i in range(5):
        col_letter = get_column_letter(5 + i)
        cell = ws.cell(row=summary_row, column=5 + i)
        cell.value = f'=COUNTIF({col_letter}5:{col_letter}{4 + len(classes)},"Done")&" / {len(classes)}"'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    widths = [14, 18, 8, 8, 16, 16, 16, 16, 16, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"Generated term template: {term_name} with {len(classes)} classes -> {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python generate_term.py <classes.json> <term_name> <output.xlsx>")
        print('Example: python generate_term.py classes.json "Term 1 2027" Term1_2027.xlsx')
        sys.exit(1)

    generate_term(sys.argv[1], sys.argv[2], sys.argv[3])
