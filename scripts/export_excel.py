"""Export data from the app's API back to an Excel spreadsheet."""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

DEFAULT_OUTPUT = r"C:\Users\Administrator.EJS\Downloads\LessonTracker_Export.xlsx"
API_BASE = "http://localhost:3000"

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
DONE_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
STARTED_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E8F0"),
)

LESSON_NAMES = [
    "Meet the Robot",
    "Power Up & Roll",
    "Robot Choreography",
    "Give It Senses",
    "The Grand Challenge",
]
LESSON_WEEKS = ["Wks 1-2", "Wks 3-4", "Wks 5-6", "Wks 7-8", "Wks 9-10"]


def export_to_excel(api_base: str = API_BASE, output_path: str = DEFAULT_OUTPUT):
    resp = requests.get(f"{api_base}/api/classes")
    resp.raise_for_status()
    data = resp.json()

    classes = data.get("classes", [])
    term = data.get("term", "Lesson Tracker")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lesson Tracker"

    # Title
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"{term} — Coding & Robotics Lesson Tracker"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center")

    # Subtitle
    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = "GigoToys S4A Robotics  |  Eshowe Junior School  |  Mr Dlamini  |  Mark each cell: Started or Done"
    sub.font = Font(size=10)
    sub.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Day", "Period", "Grade", "Class"] + [
        f"Lesson {i+1}\n({LESSON_WEEKS[i]})\n{LESSON_NAMES[i]}" for i in range(5)
    ] + ["% Complete", "Notes"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[4].height = 60

    # Data rows
    for row_idx, cls in enumerate(classes, 5):
        ws.cell(row=row_idx, column=1, value=cls["day"])
        ws.cell(row=row_idx, column=2, value=cls["period"])
        grade_cell = ws.cell(row=row_idx, column=3, value=cls["grade"])
        grade_cell.font = BOLD_FONT
        code_cell = ws.cell(row=row_idx, column=4, value=cls["classCode"])
        code_cell.font = Font(bold=True, color="1F3864")

        for i, status in enumerate(cls["lessons"]):
            cell = ws.cell(row=row_idx, column=5 + i, value=status or None)
            cell.alignment = Alignment(horizontal="center")
            if status == "Done":
                cell.fill = DONE_FILL
            elif status == "Started":
                cell.fill = STARTED_FILL

        # % Complete formula
        pct_cell = ws.cell(row=row_idx, column=10)
        pct_cell.value = f'=COUNTIF(E{row_idx}:I{row_idx},"Done")/5'
        pct_cell.number_format = "0%"
        pct_cell.font = BOLD_FONT
        pct_cell.alignment = Alignment(horizontal="center")

        # Notes
        ws.cell(row=row_idx, column=11, value=cls.get("notes", ""))

        # Borders
        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    # Summary row
    summary_row = 5 + len(classes) + 1
    ws.cell(row=summary_row, column=1, value="Classes Done — per lesson").font = BOLD_FONT
    for i in range(5):
        col_letter = get_column_letter(5 + i)
        cell = ws.cell(row=summary_row, column=5 + i)
        cell.value = f'=COUNTIF({col_letter}5:{col_letter}{4 + len(classes)},"Done")&" / {len(classes)}"'
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [14, 18, 8, 8, 16, 16, 16, 16, 16, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"Exported {len(classes)} classes to {output_path}")


if __name__ == "__main__":
    base = sys.argv[1] if len(sys.argv) > 1 else API_BASE
    out = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT
    export_to_excel(base, out)
